"""Builds small, realistic .xlsx fixtures in-memory for ingestion tests --
never committed as binary files, so diffs stay readable. Header shapes and
representative row values are taken directly from inspecting the real
Killer/Pepe sample files during Day 0, not invented.
"""

import io
from datetime import date

from openpyxl import Workbook

KILLER_HEADERS = [
    "INVOICE\nDATE",
    "NEW DATE",
    "MONTH",
    "BILL NO \nINVOICE NO",
    "PARTY NAME \n(AS PER STORE SIGNAGE)",
    "NAME",
    "STORE CODE",
    "BRAND",
    "REPORT STATUS",
    "STORE \nSTATUS",
    "TYPE",
    "ZONE",
    "CITY",
    "STATE",
    "DISTRIBUTOR \nNAME \nNEW",
    "ASM / RSM",
    "EAN CODE",
    "NEW EAN CODE",
    "MAIN\nCATEGORY",
    "ITEM NAME",
    "CATEGORY",
    "SHADE",
    "SIZE",
    "SEASON",
    "MRP",
    "FIT",
    "PRINT TYPE",
    "CLSNG \nQTY",
    "CLSNG \nVALUE",
    "QTY \nSALE",
    "NET \nSALE \nVALUE",
    "DISCOUNT \nVALUE",
    "MRP \nSALE \nVALUE",
    "F. YEAR",
]

JUNIOR_KILLER_HEADERS = [
    "INVOICE \nDATE",
    "NEW DATE",
    "MONTH",
    "BILL NO \nINVOICE NO",
    "NAME AS PER REPORT RECEIVED",
    "NAME",
    "STORE CODE",
    "BRAND",
    "REPORT \nSTATUS",
    "STORE \nSTATUS",
    "TYPE",
    "ZONE",
    "CITY",
    "STATE",
    "DISTRIBUTOR NAME",
    "ASM / RSM",
    "EAN CODE",
    "NEW EAN CODE",
    "MAIN CATEGORY",
    "ITEM NAME",
    "CATEGORY",
    "SHADE",
    "SIZE",
    "SEASON",
    "MRP",
    "FIT",
    "PRINT TYPE",
    "CLSNG \nQTY",
    "CLSNG \nVALUE",
    "QTY \nSALE",
    "NET \nSALE \nVALUE",
    "DISCOUNT \nVALUE",
    "MRP SALE\nVALUE",
]

PEPE_HEADERS = [
    "Store Name",
    "CITY",
    "STORE CODE",
    "L2L / ANULIZED",
    "Counter Types",
    "BA Name",
    "MONTH",
    "QUARTERS",
    "DATE",
    "BillNo",
    "STOCKNo",
    "PC9",
    "Size",
    "MRP",
    "Units",
    "Total MRP",
    "Net Sale Price",
    "Actual Disc",
    "WAD",
    "GENDER",
    "GEN - CAT",
    "CATEGORY",
    "FIT",
    "COLOR",
    "SEASON",
    "WEARHOUSE",
    "ATV-GWP-EOSS-Fresh",
    "Remarks (Offer/Fresh)",
]


# Real sheet names -- both real files are multi-sheet with the
# transactional data on a non-first sheet, so the seeded upload configs
# pin an exact sheet_name (validation_rules["sheet_name"]) rather than
# trusting sheet index 0. Single-sheet fixtures must use these same names
# or the pipeline (correctly) won't find the configured sheet.
KILLER_SHEET_NAME = "23-24 TO 25-26 SALE REPORT"
PEPE_SHEET_NAME = "PEPE DSR- 2026"
JUNIOR_KILLER_SHEET_NAME = "Sheet1"


def build_workbook(headers, rows, sheet_name=None):
    wb = Workbook()
    ws = wb.active
    if sheet_name:
        ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def killer_workbook(rows):
    return build_workbook(KILLER_HEADERS, rows, sheet_name=KILLER_SHEET_NAME)


def pepe_workbook(rows):
    return build_workbook(PEPE_HEADERS, rows, sheet_name=PEPE_SHEET_NAME)


def junior_killer_workbook(rows):
    return build_workbook(JUNIOR_KILLER_HEADERS, rows, sheet_name=JUNIOR_KILLER_SHEET_NAME)


def build_multi_sheet_workbook(sheets):
    """sheets: ordered list of (sheet_name, headers, rows). The first sheet
    is a decoy that becomes openpyxl's default/active sheet -- proves the
    pipeline reads the *configured* sheet_name, not sheet index 0, matching
    both real files' actual shape (summary/other sheets before the real
    transactional data sheet)."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# Real store/product identity, real category/color/season vocabulary, taken
# from the actual Killer file. MRP/NET/DISCOUNT on row 1 are set to exactly
# match the brief's mandated sanity check (MRP 2499, Net 2124, discount 375,
# ~15%), not the literal real row (which had a different discount).
KILLER_GOOD_ROWS = [
    {
        "NEW DATE": date(2023, 4, 5),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": 81,
        "NAME": "AADARSH ENTERPRISES - DUMRAO",
        "STORE CODE": "ESIS170",
        "CITY": "DUMRAO",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905646454137,
        "MAIN\nCATEGORY": "SHIRTS",
        "ITEM NAME": "1929-FS CANDY K071FSSLNDR PNK",
        "CATEGORY": "SHIRTS",
        "SHADE": "PINK",
        "SIZE": "L",
        "SEASON": "SS23",
        "MRP": 2499,
        "FIT": "KS-071 F/S SLENDER FIT",
        "PRINT TYPE": "SOLID",
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 2124,
        "DISCOUNT \nVALUE": 375,
        "MRP \nSALE \nVALUE": 2499,
        "F. YEAR": "23-24",
        "REPORT STATUS": "RECEIVED",  # always-unmapped in practice -> extra
    },
    {
        "NEW DATE": date(2023, 4, 26),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": 280,
        "NAME": "AADARSH ENTERPRISES - DUMRAO",
        "STORE CODE": "ESIS170",
        "CITY": "DUMRAO",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905533435485,
        "MAIN\nCATEGORY": "JEANS",
        "ITEM NAME": "9507-SLM- SLMFT LGHTBL",
        "CATEGORY": "JEANS",
        "SHADE": "LIGHT BLUE",
        "SIZE": 32,  # real files mix numeric jeans sizes with text sizes
        "SEASON": "AW22",
        "MRP": 3099,
        "FIT": "SLIM FIT",
        "PRINT TYPE": "(NIL)",
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 2450,
        "DISCOUNT \nVALUE": 649,
        "MRP \nSALE \nVALUE": 3099,
        "F. YEAR": "23-24",
    },
    {
        # Real return row shape: quantity/mrp/net/discount all negative,
        # unit MRP stays positive (confirmed against the real file, Day 0).
        "NEW DATE": date(2023, 4, 15),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": 417,
        "NAME": "AADARSH ENTERPRISES - DUMRAO",
        "STORE CODE": "ESIS170",
        "CITY": "DUMRAO",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905646462859,
        "MAIN\nCATEGORY": "SHIRTS",
        "ITEM NAME": "2001-FS SLOT K071FSSLNDR WT",
        "CATEGORY": "SHIRTS",
        "SHADE": "WHITE",
        "SIZE": "S",
        "SEASON": "SS23",
        "MRP": 1899,
        "FIT": "KS-071 F/S SLENDER FIT",
        "PRINT TYPE": "CHECKS",
        "QTY \nSALE": -1,
        "NET \nSALE \nVALUE": -1519,
        "DISCOUNT \nVALUE": -380,
        "MRP \nSALE \nVALUE": -1899,
        "F. YEAR": "23-24",
    },
]

# One deliberate error each: missing barcode, zero quantity, sign mismatch,
# unparseable date -- proving the error report is precise, not just "failed".
KILLER_BAD_ROWS = [
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 82,
        "NEW EAN CODE": None,  # required field missing
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 83,
        "QTY \nSALE": 0,  # zero-quantity / GWP-style row, rejected in Phase 1
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 84,
        "QTY \nSALE": -1,
        "NET \nSALE \nVALUE": -2124,
        # MRP SALE VALUE stays +2499 (inherited): negative qty + positive
        # mrp_value has no legitimate real-data match (see validation.py),
        # unlike the two now-accepted patterns -- qty>0 with mrp_value<0,
        # or a scheme discount pushing net_value <0 while mrp_value>0.
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 85,
        "NEW DATE": "not-a-date",
    },
]

# Two real sign-pattern variants found only when loading the *complete*
# real Killer file (not the ~10k-row Day 0 sample) -- both are accepted,
# not rejected, per an explicit product decision (see validation.py).
KILLER_ALT_SIGN_ROWS = [
    {
        # Pattern A: a return recorded with quantity left positive --
        # mrp_value/net_value negative is what actually signals the return.
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 501,
        "QTY \nSALE": 1,
        "MRP \nSALE \nVALUE": -2999,
        "NET \nSALE \nVALUE": -2999,
        "DISCOUNT \nVALUE": 0,
    },
    {
        # Pattern B: a normal sale where a flat scheme discount (2000)
        # exceeds the item's mrp_value (1899), pushing net_value negative
        # while mrp_value stays positive -- not a return.
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 502,
        "QTY \nSALE": 1,
        "MRP \nSALE \nVALUE": 1899,
        "DISCOUNT \nVALUE": 2000,
        "NET \nSALE \nVALUE": -101,
    },
]

# Spans 2 financial years / 3 calendar months / 3 seasons / 2 categories /
# 2 stores so trend math (YoY, MoM, Season-by-Season) and store/category
# scoping have something real to group and order by. F. YEAR is left
# consistent with each NEW DATE's fiscal year (Apr-Mar) since the MVs derive
# financial_year from dim_calendar via sale_date, not from this column
# (Killer's F. YEAR is trusted-as-supplied and only lands on fact_sales for
# audit) -- but a mismatched value here would silently mask a real bug, so
# it's kept accurate.
KILLER_TREND_ROWS = [
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 501,
        "NEW DATE": date(2023, 4, 5),
        "MONTH": "APRIL",
        "F. YEAR": "23-24",
        "CATEGORY": "SHIRTS",
        "MAIN\nCATEGORY": "SHIRTS",
        "SEASON": "SS23",
        "MRP": 1000,
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 1000,
        "DISCOUNT \nVALUE": 0,
        "MRP \nSALE \nVALUE": 1000,
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 502,
        "NEW DATE": date(2023, 7, 10),
        "MONTH": "JULY",
        "F. YEAR": "23-24",
        "CATEGORY": "SHIRTS",
        "MAIN\nCATEGORY": "SHIRTS",
        "SEASON": "SS23",
        "MRP": 2000,
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 1800,
        "DISCOUNT \nVALUE": 200,
        "MRP \nSALE \nVALUE": 2000,
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 503,
        "NEW EAN CODE": 8905533435485,
        "NEW DATE": date(2023, 10, 15),
        "MONTH": "OCTOBER",
        "F. YEAR": "23-24",
        "CATEGORY": "JEANS",
        "MAIN\nCATEGORY": "JEANS",
        "SEASON": "AW23",
        "MRP": 750,
        "QTY \nSALE": 2,
        "NET \nSALE \nVALUE": 1500,
        "DISCOUNT \nVALUE": 0,
        "MRP \nSALE \nVALUE": 1500,
    },
    {
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 504,
        "NEW DATE": date(2024, 4, 20),
        "MONTH": "APRIL",
        "F. YEAR": "24-25",
        "CATEGORY": "SHIRTS",
        "MAIN\nCATEGORY": "SHIRTS",
        "SEASON": "SS24",
        "MRP": 3000,
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 3000,
        "DISCOUNT \nVALUE": 0,
        "MRP \nSALE \nVALUE": 3000,
    },
    {
        # Different store -- proves store_trend's store_code scoping
        # actually filters, not just happens to match everything. Distinct
        # NAME too: a different store_code within one brand always means a
        # different real-world store (client-confirmed invariant -- see
        # dashboard_filter_options), so its name must differ as well or
        # store_name-based dashboard filtering couldn't tell them apart.
        **KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": 505,
        "NEW DATE": date(2023, 4, 5),
        "MONTH": "APRIL",
        "F. YEAR": "23-24",
        "STORE CODE": "ESIS999",
        "NAME": "SILVER SQUARE - PATNA",
        "CATEGORY": "SHIRTS",
        "MAIN\nCATEGORY": "SHIRTS",
        "SEASON": "SS23",
        "MRP": 500,
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 500,
        "DISCOUNT \nVALUE": 0,
        "MRP \nSALE \nVALUE": 500,
    },
]

PEPE_GOOD_ROWS = [
    {
        "Store Name": "THE BIG SHOP - PURNEA",
        "CITY": "PURNEA",
        "STORE CODE": "SI-032",
        "MONTH": "JANUARY- 2026",
        "QUARTERS": "Q- 4",
        "DATE": date(2026, 1, 15),
        "BillNo": "PRHO26-79718",
        "STOCKNo": 8905875293118,
        "PC9": "PM308998",
        "Size": "M",
        "MRP": 2999,
        "Units": 1,
        "Total MRP": 2999,
        "Net Sale Price": 1799,
        "Actual Disc": 1200,
        "WAD": 0.40,  # matches computed discount% exactly (brief's sanity check)
        "GENDER": "MENS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "SHIRTS",
        "FIT": "REGULAR",
        "COLOR": "WHITE",
        "SEASON": "FASHION BASICS",
    },
    {
        "Store Name": "THE BIG SHOP - PURNEA",
        "CITY": "PURNEA",
        "STORE CODE": "SI-032",
        "MONTH": "APRIL- 2026",  # different month -- real files span several
        "QUARTERS": "Q- 1",
        "DATE": date(2026, 4, 3),
        "BillNo": "PRHON26-8583",
        "STOCKNo": 8905875558521,
        "PC9": "PM3091063",
        "Size": "M",
        "MRP": 2799,
        "Units": 1,
        "Total MRP": 2799,
        "Net Sale Price": 2519,
        "Actual Disc": 280,
        "WAD": 0.1000357270453734,
        "GENDER": "MENS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "SHIRTS",
        "FIT": "REGULAR",
        "COLOR": "OLIVE GREEN",
        "SEASON": "AW25",
    },
    {
        # Real return row shape (Day 0 finding): all negative together.
        "Store Name": "CHANDA MAMA - HAJIPUR",
        "CITY": "HAJIPUR",
        "STORE CODE": "SI-008",
        "MONTH": "JANUARY- 2026",
        "QUARTERS": "Q- 4",
        "DATE": date(2026, 1, 24),
        "BillNo": "0101N-019614",
        "STOCKNo": 8905875451341,
        "PC9": "PM3090959",
        "Size": "XXL",
        "MRP": 3499,
        "Units": -1,
        "Total MRP": -3499,
        "Net Sale Price": -2099.03,
        "Actual Disc": -1399.97,
        "WAD": 0.4001057444984281,
        "GENDER": "MENS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "SHIRTS",
        "FIT": "REGULAR",
        "COLOR": "DARK GREEN",
        "SEASON": "AW25",
    },
]

PEPE_BAD_ROWS = [
    {
        **PEPE_GOOD_ROWS[0],
        "BillNo": "PRHO26-79719",
        "STOCKNo": None,  # required field missing
    },
    {
        **PEPE_GOOD_ROWS[0],
        "BillNo": "PRHO26-79720",
        "Units": 0,  # zero-quantity / GWP-style row
    },
    {
        **PEPE_GOOD_ROWS[0],
        "BillNo": "PRHO26-79721",
        "WAD": 0.90,  # supplied discount% wildly disagrees with computed 40%
    },
]

# Real store/product identity and header vocabulary from the actual Junior
# Killer file (kids product line, distinct brand from KILLER despite the
# near-identical column layout -- JK-prefixed store codes, BRAND="JR
# KILLER", no F. YEAR column at all).
JUNIOR_KILLER_GOOD_ROWS = [
    {
        "NEW DATE": date(2024, 4, 5),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": "0101-0022555",
        "NAME": "CHANDA MAMA - HAJIPUR",
        "STORE CODE": "JKESIS011",
        "CITY": "GAYA",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905935224182,
        "MAIN CATEGORY": "SHIRTS",
        "ITEM NAME": "001-FS GREAT KKS001FSREFIT LMN",
        "CATEGORY": "SHIRTS",
        "SHADE": "LEMON",
        "SIZE": "11-12 YEARS",
        "SEASON": "CORE",
        "MRP": 1299,
        "FIT": "KKS-001 F/S REGULAR FIT",
        "PRINT TYPE": "STRIPES",
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 1299,
        "DISCOUNT \nVALUE": 0,
        "MRP SALE\nVALUE": 1299,
        "BRAND": "JR KILLER",  # always-unmapped in practice -> extra
    },
    {
        "NEW DATE": date(2024, 4, 12),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": "0101-0011246",
        "NAME": "CHANDA MAMA - HAJIPUR",
        "STORE CODE": "JKESIS011",
        "CITY": "GAYA",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905935221884,
        "MAIN CATEGORY": "JEANS",
        "ITEM NAME": "015-FS STONE KKS001FSREFIT WT",
        "CATEGORY": "JEANS",
        "SHADE": "WHITE",
        "SIZE": "10-11 YEARS",
        "SEASON": "SS24",
        "MRP": 1599,
        "FIT": "KKS-001 F/S REGULAR FIT",
        "PRINT TYPE": "PRINT",
        "QTY \nSALE": 1,
        "NET \nSALE \nVALUE": 1299,
        "DISCOUNT \nVALUE": 300,
        "MRP SALE\nVALUE": 1599,
    },
    {
        # Return row: quantity/mrp/net/discount all negative, unit MRP
        # stays positive (same convention confirmed for Killer at Day 0).
        "NEW DATE": date(2024, 4, 20),
        "MONTH": "APRIL",
        "BILL NO \nINVOICE NO": "0101-0033012",
        "NAME": "CHANDA MAMA - HAJIPUR",
        "STORE CODE": "JKESIS011",
        "CITY": "GAYA",
        "STATE": "BIHAR",
        "ZONE": "EAST",
        "NEW EAN CODE": 8905935224182,
        "MAIN CATEGORY": "SHIRTS",
        "ITEM NAME": "001-FS GREAT KKS001FSREFIT LMN",
        "CATEGORY": "SHIRTS",
        "SHADE": "LEMON",
        "SIZE": "11-12 YEARS",
        "SEASON": "CORE",
        "MRP": 1299,
        "FIT": "KKS-001 F/S REGULAR FIT",
        "PRINT TYPE": "STRIPES",
        "QTY \nSALE": -1,
        "NET \nSALE \nVALUE": -1299,
        "DISCOUNT \nVALUE": 0,
        "MRP SALE\nVALUE": -1299,
    },
]

# One deliberate error each: missing barcode, zero quantity, sign mismatch
# (negative qty, positive mrp_value -- the one combination with no
# legitimate real-data match, see validation.py), unparseable date.
JUNIOR_KILLER_BAD_ROWS = [
    {
        **JUNIOR_KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": "0101-0099001",
        "NEW EAN CODE": None,
    },
    {
        **JUNIOR_KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": "0101-0099002",
        "QTY \nSALE": 0,
    },
    {
        **JUNIOR_KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": "0101-0099003",
        "QTY \nSALE": -1,
        "NET \nSALE \nVALUE": -1299,
        # MRP SALE VALUE stays +1299 (inherited): negative qty + positive
        # mrp_value has no legitimate real-data match.
    },
    {
        **JUNIOR_KILLER_GOOD_ROWS[0],
        "BILL NO \nINVOICE NO": "0101-0099004",
        "NEW DATE": "not-a-date",
    },
]


# Real column vocabulary/values from KRAUS - SALE REPORT-YTD-June.xlsx, the
# file the client confirmed 2026-07-18 (WhatsApp: "Use this, we will use the
# same format going forward") replaces the original one-off sample this
# brand was first onboarded against a few hours earlier the same day. Real
# file has two sheets ("REPORT", a pivot summary, and "SALE REPORT", the
# 718-row detail) -- validation_rules["sheet_name"] is "SALE REPORT",
# unlike the original mapping which had none configured at all.
KRAUS_SHEET_NAME = "SALE REPORT"
KRAUS_HEADERS = [
    "INVOICE \nDATE",
    "CORRECTED DATE",
    "MONTH",
    "INVOICE NO",
    "STORE NAME",
    "STORE CODE",
    "EAN CODE",
    "BRAND",
    "ITEM NAME",
    "CATEGORY",
    "SHADE",
    "SIZE",
    "QTY \nSALE",
    "NET \nSALE VALUE",
    "DISCOUNT \nVALUE",
    "MRP\nVALUE",
]

KRAUS_GOOD_ROWS = [
    {
        "INVOICE \nDATE": "01/01/2026",
        "CORRECTED DATE": date(2026, 1, 1),
        "MONTH": 1,
        "INVOICE NO": 11009,
        "STORE NAME": "PANKH",
        "STORE CODE": "KRA-1",
        "EAN CODE": 8905747443917,
        "BRAND": "KRAUS",
        "ITEM NAME": "LFA-2106",
        "CATEGORY": "BAGGY",
        "SHADE": "DARK BLUE",
        "SIZE": "28 / M    ",
        "QTY \nSALE": 1,
        "NET \nSALE VALUE": 1498,
        "DISCOUNT \nVALUE": 997,
        "MRP\nVALUE": 2495,
    },
    {
        # qty=2 row: proves unit_mrp is derived by dividing mrp_value by
        # quantity, not treated as already a per-unit price -- 3598 / 2 =
        # 1799, matching the same style's qty=1 per-unit price elsewhere
        # in the real file.
        "INVOICE \nDATE": date(2026, 1, 8),
        "CORRECTED DATE": date(2026, 1, 8),
        "MONTH": 1,
        "INVOICE NO": 11270,
        "STORE NAME": "PANKH",
        "STORE CODE": "KRA-1",
        "EAN CODE": 8905747513504,
        "BRAND": "KRAUS",
        "ITEM NAME": "LTT-233",
        "CATEGORY": "TREGGING",
        "SHADE": "BLACK",
        "SIZE": "32 / XL   ",
        "QTY \nSALE": 2,
        "NET \nSALE VALUE": 2160,
        "DISCOUNT \nVALUE": 1438,
        "MRP\nVALUE": 3598,
    },
    {
        # Real return row shape: quantity/net/discount/mrp all negative
        # together. STORE NAME casing ("Pankh") deliberately differs from
        # the other rows' "PANKH" -- matches a real inconsistency in the
        # client's file; store identity still resolves by STORE CODE.
        "INVOICE \nDATE": "03/02/2026",
        "CORRECTED DATE": date(2026, 2, 3),
        "MONTH": 2,
        "INVOICE NO": 12186,
        "STORE NAME": "Pankh",
        "STORE CODE": "KRA-1",
        "EAN CODE": 8905747445027,
        "BRAND": "KRAUS",
        "ITEM NAME": "LTA-2034",
        "CATEGORY": "SHIRT",
        "SHADE": "WHITE",
        "SIZE": "30 / L    ",
        "QTY \nSALE": -1,
        "NET \nSALE VALUE": -1020,
        "DISCOUNT \nVALUE": -679,
        "MRP\nVALUE": -1699,
    },
]

KRAUS_BAD_ROWS = [
    {
        **KRAUS_GOOD_ROWS[0],
        "INVOICE NO": 99001,
        "EAN CODE": None,  # required field missing
    },
    {
        **KRAUS_GOOD_ROWS[0],
        "INVOICE NO": 99002,
        "QTY \nSALE": 0,  # zero-quantity / GWP-style row
    },
    {
        **KRAUS_GOOD_ROWS[0],
        "INVOICE NO": 99003,
        "QTY \nSALE": -1,
        "NET \nSALE VALUE": -1498,
        # MRP\nVALUE stays +2495 (inherited): negative qty + positive
        # mrp_value has no legitimate real-data match.
    },
    {
        **KRAUS_GOOD_ROWS[0],
        "INVOICE NO": 99004,
        "CORRECTED DATE": "not-a-date",
    },
]


def kraus_workbook(rows):
    return build_workbook(KRAUS_HEADERS, rows, sheet_name=KRAUS_SHEET_NAME)


# Pepe Kids: client-confirmed 2026-07-18 same column template as Pepe
# menswear (PEPE_HEADERS reused as-is), but a genuinely distinct brand --
# own DimBrand row, own (brand, product_line) config, and its own store-
# code namespace even for a store that also carries Pepe menswear (same
# convention as Killer vs Junior Killer). Store codes below are
# deliberately different from PEPE_GOOD_ROWS' SI-xxx codes to reflect
# that. No real Pepe Kids sample file exists yet -- these rows mirror
# PEPE_GOOD_ROWS' shape (same column vocabulary, a normal sale, a
# different-month sale, and a return), not values taken from a real file.
PEPE_KIDS_GOOD_ROWS = [
    {
        "Store Name": "LITTLE STARS - PURNEA",
        "CITY": "PURNEA",
        "STORE CODE": "SIK-032",
        "MONTH": "JANUARY- 2026",
        "QUARTERS": "Q- 4",
        "DATE": date(2026, 1, 15),
        "BillNo": "KIDS26-10001",
        "STOCKNo": 8905875293200,
        "PC9": "PK308998",
        "Size": "5-6Y",
        "MRP": 999,
        "Units": 1,
        "Total MRP": 999,
        "Net Sale Price": 799,
        "Actual Disc": 200,
        "WAD": 0.2002002002002002,
        "GENDER": "BOYS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "T-SHIRTS",
        "FIT": "REGULAR",
        "COLOR": "BLUE",
        "SEASON": "FASHION BASICS",
    },
    {
        "Store Name": "LITTLE STARS - PURNEA",
        "CITY": "PURNEA",
        "STORE CODE": "SIK-032",
        "MONTH": "APRIL- 2026",
        "QUARTERS": "Q- 1",
        "DATE": date(2026, 4, 3),
        "BillNo": "KIDS26-10083",
        "STOCKNo": 8905875558600,
        "PC9": "PK3091063",
        "Size": "7-8Y",
        "MRP": 1299,
        "Units": 1,
        "Total MRP": 1299,
        "Net Sale Price": 1169,
        "Actual Disc": 130,
        "WAD": 0.10007698229407236,
        "GENDER": "GIRLS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "FROCKS",
        "FIT": "REGULAR",
        "COLOR": "PINK",
        "SEASON": "AW25",
    },
    {
        # Real return row shape (same convention as Pepe menswear): all
        # negative together.
        "Store Name": "KIDDIE CORNER - HAJIPUR",
        "CITY": "HAJIPUR",
        "STORE CODE": "SIK-008",
        "MONTH": "JANUARY- 2026",
        "QUARTERS": "Q- 4",
        "DATE": date(2026, 1, 24),
        "BillNo": "0101K-019700",
        "STOCKNo": 8905875451400,
        "PC9": "PK3090959",
        "Size": "3-4Y",
        "MRP": 899,
        "Units": -1,
        "Total MRP": -899,
        "Net Sale Price": -539.4,
        "Actual Disc": -359.6,
        "WAD": 0.40,
        "GENDER": "BOYS",
        "GEN - CAT": "TOP WEAR",
        "CATEGORY": "T-SHIRTS",
        "FIT": "REGULAR",
        "COLOR": "GREEN",
        "SEASON": "AW25",
    },
]

PEPE_KIDS_BAD_ROWS = [
    {
        **PEPE_KIDS_GOOD_ROWS[0],
        "BillNo": "KIDS26-90001",
        "STOCKNo": None,  # required field missing
    },
    {
        **PEPE_KIDS_GOOD_ROWS[0],
        "BillNo": "KIDS26-90002",
        "Units": 0,  # zero-quantity / GWP-style row
    },
    {
        **PEPE_KIDS_GOOD_ROWS[0],
        "BillNo": "KIDS26-90003",
        "Units": -1,
        "Net Sale Price": -799,
        # Total MRP stays +999 (inherited): negative qty + positive
        # mrp_value has no legitimate real-data match.
    },
    {
        **PEPE_KIDS_GOOD_ROWS[0],
        "BillNo": "KIDS26-90004",
        "DATE": "not-a-date",
    },
]


def pepe_kids_workbook(rows):
    return build_workbook(PEPE_HEADERS, rows, sheet_name=PEPE_SHEET_NAME)
